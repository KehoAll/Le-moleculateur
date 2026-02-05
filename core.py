import os
import re
import ast
import math

import numpy as np
from openpyxl import load_workbook


def extract_molar_masses_and_uncertainties(path, sheet_name=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Fichier Excel introuvable: {path}")
    wb = load_workbook(path, data_only=True)
    try:
        if sheet_name is None:
            ws = wb.active
        elif isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]
    except (IndexError, KeyError):
        wb.close()
        raise ValueError(f"Feuille Excel introuvable: {sheet_name!r}") from None

    header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    def norm_header(value):
        if value is None:
            return None
        return str(value).strip()

    headers = [norm_header(value) for value in header_row]

    def find_header(target):
        target_norm = target.lower()
        for idx, value in enumerate(headers):
            if value is None:
                continue
            if value.lower() == target_norm:
                return idx
        return None

    symbol_idx = find_header("Symbol")
    weight_idx = find_header("Standard atomic weight")

    uncertainty_idx = find_header("Unnamed: 4")
    if uncertainty_idx is None:
        for idx, value in enumerate(headers):
            if value and "uncertainty" in value.lower():
                uncertainty_idx = idx
                break
    if uncertainty_idx is None:
        row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
        for idx, value in enumerate(row2):
            if value and "uncertainty" in str(value).lower():
                uncertainty_idx = idx
                break
    if uncertainty_idx is None and weight_idx is not None:
        if weight_idx + 1 < len(headers):
            uncertainty_idx = weight_idx + 1

    missing_columns = []
    if symbol_idx is None:
        missing_columns.append("Symbol")
    if weight_idx is None:
        missing_columns.append("Standard atomic weight")
    if uncertainty_idx is None:
        missing_columns.append("Uncertainty")
    if missing_columns:
        wb.close()
        raise ValueError(
            "Colonnes manquantes dans le fichier Excel: "
            + ", ".join(missing_columns)
        )

    def parse_maybe_interval(value):
        if isinstance(value, str):
            try:
                pair = ast.literal_eval(value)  # ex: "[10.8, 11.0]"
                return float(np.mean(pair))
            except (ValueError, SyntaxError):
                pass
        return float(value)

    def is_missing(value):
        if value is None:
            return True
        if isinstance(value, str):
            return value.strip() == ""
        if isinstance(value, float) and math.isnan(value):
            return True
        return False

    result = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        symbol = row[symbol_idx] if symbol_idx < len(row) else None
        if symbol is None:
            continue
        symbol = str(symbol).strip()
        if not symbol:
            continue

        mass = row[weight_idx] if weight_idx < len(row) else None
        err = row[uncertainty_idx] if uncertainty_idx < len(row) else None

        if mass is None:
            continue

        m_val = parse_maybe_interval(mass)

        if is_missing(err):
            e_val = 0.0
        else:
            try:
                pair_err = ast.literal_eval(str(err))
                e_val = abs(pair_err[1] - pair_err[0]) / 2.0
            except (ValueError, SyntaxError, TypeError):
                e_val = float(err)

        result[symbol] = (m_val, e_val)

    wb.close()
    return result


def parse_formula(formula):
    formula = re.sub(r"\s+", "", formula)
    if not formula:
        raise ValueError("Formule brute vide ou invalide.")

    def strip_charge(fragment):
        fragment = re.sub(r"\^[0-9]+[+-]$", "", fragment)
        fragment = re.sub(r"[+-]$", "", fragment)
        return fragment

    def parse_fragment(fragment):
        token_pattern = r"([A-Z][a-z]?|\d+(?:\.\d+)?|\(|\))"
        tokens = re.findall(token_pattern, fragment)
        if not tokens:
            raise ValueError("Formule brute vide ou invalide.")

        def is_number(token):
            return re.fullmatch(r"\d+(?:\.\d+)?", token) is not None

        def is_element(token):
            return re.fullmatch(r"[A-Z][a-z]?", token) is not None

        stack = [{}]
        i = 0
        while i < len(tokens):
            tok = tokens[i]
            if tok == "(":
                stack.append({})
                i += 1
                continue
            if tok == ")":
                if len(stack) == 1:
                    raise ValueError("Parenthese fermante sans ouvrante.")
                group = stack.pop()
                multiplier = 1.0
                if i + 1 < len(tokens) and is_number(tokens[i + 1]):
                    multiplier = float(tokens[i + 1])
                    i += 1
                for el, count in group.items():
                    stack[-1][el] = stack[-1].get(el, 0.0) + count * multiplier
                i += 1
                continue
            if is_element(tok):
                multiplier = 1.0
                if i + 1 < len(tokens) and is_number(tokens[i + 1]):
                    multiplier = float(tokens[i + 1])
                    i += 1
                stack[-1][tok] = stack[-1].get(tok, 0.0) + multiplier
                i += 1
                continue
            if is_number(tok):
                raise ValueError(f"Nombre inattendu dans la formule: '{tok}'.")
            raise ValueError(f"Jeton inconnu dans la formule: '{tok}'.")

        if len(stack) != 1:
            raise ValueError("Parenthese ouvrante sans fermante.")

        return stack[0]

    parts = re.split(r"[·.]", formula)
    total = {}
    for raw in parts:
        fragment = strip_charge(raw)
        if not fragment:
            continue
        coeff_match = re.match(r"^(\d+(?:\.\d+)?)(.*)$", fragment)
        if coeff_match:
            part_coeff = float(coeff_match.group(1))
            fragment = coeff_match.group(2)
        else:
            part_coeff = 1.0
        if not fragment:
            raise ValueError("Formule brute invalide apres coefficient.")
        fragment_dict = parse_fragment(fragment)
        for el, count in fragment_dict.items():
            total[el] = total.get(el, 0.0) + count * part_coeff
    return total


COMMON_SPECTATOR_FORMULAS = [
    "Cl",
    "Br",
    "I",
    "F",
    "NO3",
    "SO4",
    "CO3",
    "PO4",
    "OH",
    "NH4",
    "ClO4",
    "Na",
    "K",
    "Li",
    "Mg",
    "Ca",
    "Ba",
    "Sr",
    "Ag",
    "Cu",
    "Zn",
    "Fe",
    "Al",
    "Mn",
    "Co",
    "Ni",
    "Pb",
]


def _can_remove_spectator(prec_dict, spec_dict, eps=1e-12):
    k_max = None
    for el, count in spec_dict.items():
        if count <= 0:
            continue
        available = prec_dict.get(el, 0.0)
        k_el = math.floor((available + eps) / count)
        if k_max is None or k_el < k_max:
            k_max = k_el
    return k_max is not None and k_max > 0


def suggest_spectators(solute_dict, precursor_dicts, candidates=None):
    if candidates is None:
        candidates = COMMON_SPECTATOR_FORMULAS

    solute_elements = set(solute_dict.keys())
    suggestions = []
    seen = set()

    for name in candidates:
        if name in seen:
            continue
        try:
            spec_dict = parse_formula(name)
        except Exception:
            continue
        key_elements = {el for el in spec_dict.keys() if el not in ("O", "H")}
        if not key_elements:
            continue
        if key_elements.issubset(solute_elements):
            continue
        if any(_can_remove_spectator(prec_dict, spec_dict) for prec_dict in precursor_dicts):
            suggestions.append((name, spec_dict))
            seen.add(name)

    return suggestions


def parse_spectator_list(raw_list):
    spectators = []
    if not raw_list:
        return spectators
    tokens = re.split(r"[,\s]+", raw_list.strip())
    for tok in tokens:
        tok = tok.strip()
        if not tok:
            continue
        try:
            d_tok = parse_formula(tok)
            if not d_tok:
                print(f"Avertissement : ion spectateur '{tok}' non reconnu, ignore.")
                continue
            spectators.append((tok, d_tok))
        except Exception as ex:
            print(f"Avertissement : ion spectateur '{tok}' non interpretable ({ex}), ignore.")
    return spectators


def remove_spectators(prec_dict, spectators, eps=1e-12):
    active = dict(prec_dict)
    removed = []
    for spec_name, spec_dict in spectators:
        if not spec_dict:
            continue
        k_max = None
        for el, count in spec_dict.items():
            if count <= 0:
                continue
            available = active.get(el, 0.0)
            k_el = math.floor((available + eps) / count)
            if k_max is None or k_el < k_max:
                k_max = k_el
        if k_max is None:
            k_max = 0
        if k_max > 0:
            for el, count in spec_dict.items():
                active[el] = active.get(el, 0.0) - k_max * count
                if abs(active[el]) <= eps:
                    active.pop(el, None)
            removed.append((spec_name, int(k_max)))
    if not active:
        print("Avertissement : precurseur compose uniquement d'ions spectateurs apres retrait.")
    return active, removed


def validate_elements(formula_dict, dict_masses, label):
    missing = [el for el in formula_dict.keys() if el not in dict_masses]
    if missing:
        raise ValueError(
            f"Elements inconnus dans {label}: {', '.join(sorted(missing))}"
        )


def get_molar_mass_and_uncert(comp_dict, dict_masses):
    M = 0.0
    dM2 = 0.0
    for el, count in comp_dict.items():
        if el not in dict_masses:
            raise ValueError(f"Element '{el}' introuvable dans le tableau de masses atomiques.")
        M_el, dM_el = dict_masses[el]
        M += count * M_el
        dM2 += (count * dM_el) ** 2
    return M, np.sqrt(dM2)


def nnls_solve(P, f, tol=1e-12, max_iter=None):
    if max_iter is None:
        max_iter = 3 * P.shape[1]

    x = np.zeros(P.shape[1])
    passive = np.zeros(P.shape[1], dtype=bool)
    w = P.T @ (f - P @ x)
    iteration = 0

    while np.any(w > tol) and iteration < max_iter:
        t = int(np.argmax(w))
        passive[t] = True
        z = np.zeros_like(x)
        while True:
            if not np.any(passive):
                break
            P_passive = P[:, passive]
            z_passive, _, _, _ = np.linalg.lstsq(P_passive, f, rcond=None)
            z[passive] = z_passive
            z[~passive] = 0.0
            if np.all(z[passive] > tol):
                break
            negative = (z <= tol) & passive
            alpha = np.min(x[negative] / (x[negative] - z[negative]))
            x = x + alpha * (z - x)
            newly_inactive = (x <= tol) & passive
            passive[newly_inactive] = False
            z[newly_inactive] = 0.0
        x = z
        w = P.T @ (f - P @ x)
        iteration += 1

    return x


def solve_stoichiometry(final_dict, precursor_dicts, nonneg=False):
    elements = list(final_dict.keys())
    nb_el = len(elements)
    nb_prec = len(precursor_dicts)

    P = np.zeros((nb_el, nb_prec))
    f = np.zeros(nb_el)

    for i, el in enumerate(elements):
        f[i] = final_dict[el]
        for j, pdict in enumerate(precursor_dicts):
            P[i, j] = pdict.get(el, 0.0)

    if nonneg:
        x = nnls_solve(P, f)
        residual = np.sum((P @ x - f) ** 2)
        residuals = np.array([residual])
    else:
        x, residuals, _, _ = np.linalg.lstsq(P, f, rcond=None)

    rank = np.linalg.matrix_rank(P)
    singular_values = np.linalg.svd(P, compute_uv=False)
    return x, residuals, rank, singular_values
